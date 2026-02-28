"""
ICEBERG OFFICE — Server
========================
Запуск:     python server.py
Установка:  pip install flask flask-cors requests beautifulsoup4 openpyxl
"""
import os, io, threading, datetime
from flask import Flask, jsonify, request, Response
from flask_cors import CORS
from scraper import (
    init_db, scrape_firmy_cz, save_contacts, save_session,
    get_all_contacts, get_stats, update_contact, delete_contact,
    clear_all, CATEGORIES, CITIES, DB_FILE, build_url
)

app = Flask(__name__)
CORS(app)

state = {
    "running":  False,
    "progress": 0,
    "status":   "Ожидание",
    "log":      [],
    "stop":     False,
}


def log(msg, t="info"):
    state["log"].append({
        "time":  datetime.datetime.now().strftime("%H:%M:%S"),
        "msg":   msg,
        "level": t,
    })
    print(f"[{state['log'][-1]['time']}] {msg}")


def run_scrape(params):
    state.update(running=True, stop=False, progress=0, log=[])

    category_slug = params.get("category", "Restauracni-a-pohostinske-sluzby")
    city_key      = params.get("city", "Praha")
    category_name = CATEGORIES.get(category_slug, category_slug)

    log("🚀 Сбор запущен", "info")
    log(f"📂 Категория: {category_name}", "info")
    log(f"📍 Город: {city_key}", "info")

    # Показываем сгенерированный URL для отладки
    test_url = build_url(category_slug, city_key, page=1)
    log(f"🔗 URL: {test_url}", "info")

    def upd(msg, t="info"):
        log(msg, t)
        state["progress"] = min(95, max(5, len(state["log"]) * 3))
        state["status"] = msg[:70] if msg else "..."

    try:
        results = scrape_firmy_cz(
            category_slug=category_slug,
            city_key=city_key,
            log_fn=upd,
            stop_flag=lambda: state["stop"],
        )

        state["progress"] = 96
        log("💾 Сохранение в базу...", "info")

        saved  = save_contacts(results, city=city_key, category=category_name)
        emails = sum(1 for r in results if r.get("email"))
        phones = sum(1 for r in results if r.get("phone"))

        save_session(
            dt           = datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
            category     = category_name,
            city         = city_key,
            total_found  = len(results),
            total_saved  = saved,
            emails       = emails,
            phones       = phones,
        )

        log("✅ ГОТОВО!", "ok")
        log(f"   Найдено всего:   {len(results)}", "ok")
        log(f"   Новых сохранено: {saved}", "ok")
        log(f"   С email:         {emails}", "ok")
        log(f"   С телефоном:     {phones}", "ok")

    except Exception as e:
        log(f"❌ Ошибка: {e}", "err")
        import traceback
        log(traceback.format_exc(), "err")

    state["progress"] = 100
    state["running"]  = False
    state["status"]   = "Готово"


# ── API ────────────────────────────────────────────────────────────

@app.route("/api/status")
def api_status():
    return jsonify({
        "running":  state["running"],
        "progress": state["progress"],
        "status":   state["status"],
        "db":       DB_FILE,
    })

@app.route("/api/categories")
def api_categories():
    # Возвращаем список {slug, name} для дашборда
    return jsonify([{"slug": k, "name": v} for k, v in CATEGORIES.items()])

@app.route("/api/cities")
def api_cities():
    return jsonify(list(CITIES.keys()))

@app.route("/api/start", methods=["POST"])
def api_start():
    if state["running"]:
        return jsonify({"error": "Уже запущен"}), 400
    params = request.json or {}
    threading.Thread(target=run_scrape, args=(params,), daemon=True).start()
    return jsonify({"ok": True})

@app.route("/api/stop", methods=["POST"])
def api_stop():
    state["stop"]    = True
    state["running"] = False
    log("⛔ Остановлено", "err")
    return jsonify({"ok": True})

@app.route("/api/progress")
def api_progress():
    return jsonify({
        "progress": state["progress"],
        "running":  state["running"],
        "status":   state["status"],
    })

@app.route("/api/log")
def api_log():
    since = int(request.args.get("since", 0))
    return jsonify(state["log"][since:])

@app.route("/api/contacts")
def api_contacts():
    return jsonify(get_all_contacts(
        city     = request.args.get("city")     or None,
        category = request.args.get("category") or None,
        status   = request.args.get("status")   or None,
        search   = request.args.get("search")   or None,
        sort     = request.args.get("sort", "date"),
        limit    = int(request.args.get("limit",  5000)),
        offset   = int(request.args.get("offset", 0)),
    ))

@app.route("/api/contacts/<int:cid>", methods=["PUT"])
def api_update(cid):
    update_contact(cid, request.json or {})
    return jsonify({"ok": True})

@app.route("/api/contacts/<int:cid>", methods=["DELETE"])
def api_delete(cid):
    delete_contact(cid)
    return jsonify({"ok": True})

@app.route("/api/stats")
def api_stats():
    return jsonify(get_stats())

@app.route("/api/db/clear", methods=["POST"])
def api_clear():
    clear_all()
    return jsonify({"ok": True})


# ── EXPORT ─────────────────────────────────────────────────────────

@app.route("/api/export/csv")
def api_csv():
    contacts = get_all_contacts(
        city     = request.args.get("city")     or None,
        category = request.args.get("category") or None,
    )
    cols  = ["id","name","category","city","address","phone","email","website","description","status","quality","date_added"]
    heads = ["#","Название","Категория","Город","Адрес","Телефон","Email","Сайт","Описание","Статус","Качество","Дата"]
    rows  = [heads] + [[str(c.get(k,"") or "") for k in cols] for c in contacts]
    csv   = "\n".join([",".join([f'"{v.replace(chr(34),chr(34)*2)}"' for v in row]) for row in rows])
    return Response(
        "\ufeff" + csv,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=iceberg_contacts.csv"}
    )

@app.route("/api/export/email-list")
def api_email_list():
    contacts = get_all_contacts(
        city     = request.args.get("city")     or None,
        category = request.args.get("category") or None,
    )
    with_email = [c for c in contacts if c.get("email")]
    rows = [["Email","Название","Город","Категория","Телефон","Сайт"]]
    rows += [[c.get("email",""), c.get("name",""), c.get("city",""),
              c.get("category",""), c.get("phone",""), c.get("website","")] for c in with_email]
    csv = "\n".join([",".join([f'"{v}"' for v in r]) for r in rows])
    return Response(
        "\ufeff" + csv,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=email_list.csv"}
    )

@app.route("/api/export/xlsx")
def api_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "pip install openpyxl"}), 500

    contacts = get_all_contacts(
        city     = request.args.get("city")     or None,
        category = request.args.get("category") or None,
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ICEBERG Contacts"

    cols   = ["id","name","category","city","address","phone","email","website","description","status","quality","date_added"]
    heads  = ["#","Название","Категория","Город","Адрес","Телефон","Email","Сайт","Описание","Статус","Кач.%","Дата"]
    widths = [5, 36, 24, 16, 28, 18, 30, 32, 40, 14, 8, 14]

    hfill = PatternFill("solid", fgColor="0C0C0B")
    hfont = Font(bold=True, color="2D7AEE", name="Calibri", size=11)
    for i, (h, w) in enumerate(zip(heads, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}1"

    for ri, c in enumerate(contacts, 2):
        bg = PatternFill("solid", fgColor="161614" if ri % 2 == 0 else "1A1A18")
        for ci, k in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=c.get(k,"") or "")
            cell.fill      = bg
            cell.font      = Font(name="Calibri", size=10, color="D4D2CA")
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[ri].height = 17

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=iceberg_contacts.xlsx"}
    )


@app.route("/")
def index():
    s = get_stats()
    return (
        f'<html><body style="background:#0c0c0b;color:#d4d2ca;font-family:monospace;padding:40px">'
        f'<h2 style="color:#2d7aee">✅ ICEBERG OFFICE — Server Running</h2>'
        f'<p>База: <b style="color:#2d7aee">{s["total"]} контактов</b> | '
        f'Email: {s["with_email"]} | Телефон: {s["with_phone"]} | Сайты: {s["with_website"]}</p>'
        f'<p style="color:#4a4a46">DB: {DB_FILE}</p>'
        f'</body></html>'
    )


if __name__ == "__main__":
    init_db()
    PORT = int(os.environ.get("PORT", 5000))
    print(f"\n{'='*46}")
    print(f"  ICEBERG OFFICE — Server v2")
    print(f"  http://localhost:{PORT}")
    print(f"  DB: {DB_FILE}")
    print(f"{'='*46}\n")
    app.run(host="0.0.0.0", port=PORT, debug=False, threaded=True)
